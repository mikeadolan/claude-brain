#!/usr/bin/env python3
"""
brain_tag_review.py - Batch tag review for brain sessions.

Generates an xlsx of untagged (or all) sessions with auto-suggested tags.
User reviews and edits tags in the spreadsheet, then runs update mode
to write tags back to the database.

Two modes:
  --generate     Create xlsx with suggested tags for review
  --update       Read reviewed xlsx and update tags in DB

Options:
  --all          Include already-tagged sessions (default: untagged only)
  --project <p>  Filter to a specific project prefix
  --output <f>   Output path for xlsx (default: tag_review.xlsx)

Workflow:
  1. python3 scripts/brain_tag_review.py --generate
  2. Review tag_review.xlsx -- edit the yellow tags column
  3. python3 scripts/brain_tag_review.py --update --map tag_review.xlsx

Usage via slash command: /brain-tag-review
"""

import argparse
import os
import re
import sqlite3
import sys
from datetime import datetime, timezone
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
ROOT_DIR = SCRIPT_DIR.parent

# Same tag keywords as import scripts
TAG_KEYWORDS = {
    "book-editing": ["edit", "polish", "chapter", "manuscript", "copyedit", "proofread", "rewrite", "revision"],
    "memoir": ["memoir", "johnny", "goods", "maffia", "mob", "harlem", "gangster", "three fingers", "pleasant avenue"],
    "dialogue": ["dialogue", "dialog", "conversation", "speech", "quotes"],
    "job-search": ["resume", "job", "interview", "career", "hiring", "salary", "recruiter", "linkedin"],
    "finance": ["money", "bank", "invest", "stock", "crypto", "budget", "tax", "financial"],
    "coding": ["python", "script", "code", "programming", "debug", "function", "api", "github"],
    "ai-tools": ["chatgpt", "claude", "gpt", "openai", "anthropic", "gemini", "ai model", "llm", "prompt"],
    "tech-setup": ["linux", "fedora", "install", "setup", "config", "terminal", "laptop", "asus", "ubuntu", "samba"],
    "family": ["mom", "mother", "father", "wife", "daughter", "son", "family", "sister", "brother", "grandpa"],
    "health": ["doctor", "medical", "health", "therapy", "leg", "injury", "exercise", "weight"],
    "legal": ["lawyer", "legal", "court", "contract", "lawsuit", "attorney", "will", "estate"],
    "home": ["house", "home", "repair", "plumbing", "electric", "hvac", "mortgage", "renovation"],
    "auto": ["car", "toyota", "highlander", "vehicle", "mechanic", "tire", "oil"],
    "travel": ["travel", "flight", "hotel", "trip", "vacation", "passport", "airport"],
    "music": ["music", "song", "guitar", "band", "album", "spotify"],
    "business": ["business", "company", "startup", "entrepreneur", "marketing", "client", "consultant"],
    "writing": ["writing", "write", "story", "narrative", "voice", "author", "publisher"],
    "research": ["research", "compare", "analysis", "review", "best", "recommend", "options"],
    "brain-project": ["brain", "mcp", "hook", "session", "memory", "database", "sqlite", "ingestion"],
}


def green(text):
    return f"\033[92m{text}\033[0m"

def yellow(text):
    return f"\033[93m{text}\033[0m"

def red(text):
    return f"\033[91m{text}\033[0m"


def load_config():
    config_path = ROOT_DIR / "config.yaml"
    if not config_path.exists():
        print(red("config.yaml not found."))
        sys.exit(1)
    import yaml
    with open(config_path) as f:
        config = yaml.safe_load(f)
    storage = config.get("storage", {})
    mode = storage.get("mode", "local")
    if mode == "synced":
        return os.path.expanduser(storage.get("local_db_path", ""))
    else:
        return os.path.join(os.path.expanduser(storage.get("root_path", "")), "claude-brain.db")


def suggest_tags(text):
    if not text:
        return ""
    text_lower = text.lower()
    matched = []
    for tag, keywords in TAG_KEYWORDS.items():
        if any(kw in text_lower for kw in keywords):
            matched.append(tag)
    return ", ".join(matched[:3]) if matched else ""


def cmd_generate(args):
    """Generate xlsx with sessions and suggested tags."""
    db_path = load_config()
    conn = sqlite3.connect(db_path)

    # Build query
    if args.all:
        where = "WHERE 1=1"
    else:
        where = "WHERE (tags IS NULL OR tags = '')"

    if args.project:
        where += f" AND project = '{args.project}'"

    rows = conn.execute(f"""
        SELECT session_id, project, started_at, source, message_count, tags, notes
        FROM sys_sessions
        {where}
        ORDER BY started_at
    """).fetchall()

    print(f"\nSessions to review: {len(rows)}")

    if not rows:
        print("No untagged sessions found. Use --all to include already-tagged sessions.")
        conn.close()
        return

    # Build xlsx
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    output_path = Path(args.output) if args.output else ROOT_DIR / "tag_review.xlsx"
    wb = Workbook()

    # Tab 1: Sessions
    ws1 = wb.active
    ws1.title = "Sessions"
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    edit_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    headers = ["session_id", "project", "date", "source", "messages", "current_tags", "suggested_tags", "summary"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row_num, r in enumerate(rows, 2):
        session_id, project, started_at, source, msg_count, current_tags, notes = r
        date = started_at[:10] if started_at else ""

        # Auto-suggest tags from notes content
        suggested = suggest_tags(notes or "")

        ws1.cell(row=row_num, column=1, value=session_id)
        ws1.cell(row=row_num, column=2, value=project)
        ws1.cell(row=row_num, column=3, value=date)
        ws1.cell(row=row_num, column=4, value=source or "")
        ws1.cell(row=row_num, column=5, value=msg_count)
        ws1.cell(row=row_num, column=6, value=current_tags or "")
        cell_tags = ws1.cell(row=row_num, column=7, value=suggested)
        cell_tags.fill = edit_fill
        ws1.cell(row=row_num, column=8, value=(notes or "")[:150])

    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 8
    ws1.column_dimensions["C"].width = 12
    ws1.column_dimensions["D"].width = 15
    ws1.column_dimensions["E"].width = 10
    ws1.column_dimensions["F"].width = 25
    ws1.column_dimensions["G"].width = 30
    ws1.column_dimensions["H"].width = 60

    # Tab 2: Tag Reference
    ws2 = wb.create_sheet("Tag Reference")
    ws2.cell(row=1, column=1, value="AVAILABLE TAGS").font = Font(bold=True, size=14)
    tag_headers = ["tag", "description"]
    for col, h in enumerate(tag_headers, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    tag_descriptions = {
        "auto": "Cars, vehicles, maintenance", "ai-tools": "ChatGPT, Claude, Gemini, LLMs",
        "book-editing": "Chapter edits, polish, revisions", "brain-project": "Brain DB, MCP, hooks development",
        "business": "Companies, startups, marketing", "coding": "Python, scripts, programming",
        "dialogue": "Dialogue writing in the book", "family": "Mom, father, wife, kids, grandpa",
        "finance": "Money, banking, investing, taxes", "health": "Medical, therapy, injuries",
        "home": "House repairs, plumbing, electrical", "job-search": "Resumes, interviews, career",
        "legal": "Lawyers, courts, contracts, estates", "memoir": "Johnny Goods, mob stories, Harlem",
        "music": "Songs, guitar, music project", "research": "Comparisons, analysis, reviews",
        "tech-setup": "Linux, Fedora, installs, configs", "travel": "Flights, hotels, trips",
        "writing": "Narrative craft, voice, storytelling",
    }
    for row_num, (tag, desc) in enumerate(sorted(tag_descriptions.items()), 4):
        ws2.cell(row=row_num, column=1, value=tag)
        ws2.cell(row=row_num, column=2, value=desc)

    last_row = 4 + len(tag_descriptions)
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 50
    ws2.cell(row=last_row + 1, column=1, value="").font = Font(bold=True)
    ws2.cell(row=last_row + 2, column=1, value="INSTRUCTIONS:").font = Font(bold=True)
    ws2.cell(row=last_row + 3, column=1, value="1. Edit the yellow SUGGESTED_TAGS column on the Sessions tab")
    ws2.cell(row=last_row + 4, column=1, value="2. You can use any tag from this list or create new ones")
    ws2.cell(row=last_row + 5, column=1, value="3. Multiple tags: separate with commas")
    ws2.cell(row=last_row + 6, column=1, value="4. Leave blank to skip tagging a session")
    ws2.cell(row=last_row + 7, column=1, value="5. Save the file, then run: python3 scripts/brain_tag_review.py --update --map tag_review.xlsx")

    wb.save(output_path)
    conn.close()

    # Summary
    print(f"\n{green('Tag review xlsx generated:')} {output_path}")
    print(f"\n  Sessions: {len(rows)}")
    if args.project:
        print(f"  Project filter: {args.project}")
    print(f"  Mode: {'all sessions' if args.all else 'untagged only'}")
    print(f"\n{yellow('NEXT:')} Review the yellow suggested_tags column, then run:")
    print(f"  python3 scripts/brain_tag_review.py --update --map {output_path}")


def cmd_update(args):
    """Read reviewed xlsx and update tags in DB."""
    if not args.map:
        print(red("--map <xlsx_file> is required for update mode"))
        sys.exit(1)

    map_path = Path(args.map)
    if not map_path.exists():
        print(red(f"File not found: {map_path}"))
        sys.exit(1)

    from openpyxl import load_workbook
    wb = load_workbook(map_path)
    ws = wb["Sessions"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    sid_col = headers.index("session_id") + 1
    tags_col = headers.index("suggested_tags") + 1

    updates = []
    for row in range(2, ws.max_row + 1):
        sid = ws.cell(row=row, column=sid_col).value
        tags = ws.cell(row=row, column=tags_col).value or ""
        if sid and tags.strip():
            updates.append((tags.strip(), sid))

    print(f"\nSessions with tags to update: {len(updates)}")

    if not updates:
        print("No tags to update.")
        return

    db_path = load_config()
    conn = sqlite3.connect(db_path)

    updated = 0
    for tags, sid in updates:
        conn.execute("UPDATE sys_sessions SET tags = ? WHERE session_id = ?", (tags, sid))
        updated += 1

    conn.commit()
    conn.close()

    print(f"\n{green('Tags updated:')} {updated} sessions")


def main():
    parser = argparse.ArgumentParser(description="Batch tag review for brain sessions")
    parser.add_argument("--generate", action="store_true", help="Generate xlsx with suggested tags")
    parser.add_argument("--update", action="store_true", help="Update DB from reviewed xlsx")
    parser.add_argument("--map", metavar="XLSX", help="Reviewed xlsx file (for --update)")
    parser.add_argument("--all", action="store_true", help="Include already-tagged sessions")
    parser.add_argument("--project", metavar="PREFIX", help="Filter to one project prefix")
    parser.add_argument("--output", metavar="FILE", help="Output xlsx path")

    args = parser.parse_args()

    if args.generate:
        cmd_generate(args)
    elif args.update:
        cmd_update(args)
    else:
        parser.print_help()
        print(f"\n{yellow('Workflow:')}")
        print(f"  1. python3 scripts/brain_tag_review.py --generate")
        print(f"  2. Review tag_review.xlsx (edit suggested_tags column)")
        print(f"  3. python3 scripts/brain_tag_review.py --update --map tag_review.xlsx")


if __name__ == "__main__":
    main()
