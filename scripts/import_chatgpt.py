#!/usr/bin/env python3
"""
import_chatgpt.py - Import ChatGPT conversation exports into the brain database.

Supports the full data export from OpenAI (Settings → Data → Export).
Text conversations only -- audio, images, and system messages are skipped.

Three modes:
  --scan <dir>     Generate mapping CSV with suggested project assignments
  --import <dir>   Import conversations using reviewed mapping CSV
  --dry-run        Preview import without touching the database

Workflow:
  1. Export from ChatGPT (Settings → Data → Export → wait for email → download zip)
  2. Extract the zip to a folder
  3. Run: python3 scripts/import_chatgpt.py --scan imports/chatgpt-export/
  4. Review chatgpt_import_map.csv -- edit project assignments, delete unwanted rows
  5. Run: python3 scripts/import_chatgpt.py --import imports/chatgpt-export/ --map chatgpt_import_map.csv --dry-run
  6. When satisfied: python3 scripts/import_chatgpt.py --import imports/chatgpt-export/ --map chatgpt_import_map.csv
"""

import argparse
import csv
import json
import os
import sqlite3
import sys
from datetime import datetime, timezone
from pathlib import Path

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
ROOT_DIR = SCRIPT_DIR.parent

# Project keyword suggestions for auto-tagging
PROJECT_KEYWORDS = {
    "jg": ["johnny", "goods", "memoir", "chapter", "ch1", "ch2", "ch3", "ch4",
            "ch5", "ch6", "ch7", "ch8", "ch9", "ch10", "ch11", "polish",
            "manuscript", "mob", "harlem", "mafia", "maffia", "three fingers",
            "gangster", "dialogue", "copyedit", "proofread", "book edit",
            "edit instructions", "summary deliverable"],
    "js": ["resume", "job", "interview", "career", "hiring", "salary",
           "linkedin", "cover letter", "recruiter"],
}
DEFAULT_PROJECT = "gen"

# Tag keyword suggestions for auto-tagging (broad topics, ~20-30 total)
TAG_KEYWORDS = {
    "book-editing": ["edit", "polish", "chapter", "manuscript", "copyedit", "proofread", "rewrite", "revision"],
    "memoir": ["memoir", "johnny", "goods", "maffia", "mob", "harlem", "gangster", "three fingers"],
    "dialogue": ["dialogue", "dialog", "conversation", "speech", "quotes"],
    "job-search": ["resume", "job", "interview", "career", "hiring", "salary", "recruiter", "linkedin"],
    "finance": ["money", "bank", "invest", "stock", "crypto", "budget", "tax", "financial"],
    "coding": ["python", "script", "code", "programming", "debug", "function", "api", "github"],
    "ai-tools": ["chatgpt", "claude", "gpt", "openai", "anthropic", "gemini", "ai model", "llm", "prompt"],
    "tech-setup": ["linux", "fedora", "install", "setup", "config", "terminal", "laptop", "asus"],
    "family": ["mom", "mother", "father", "wife", "daughter", "son", "family", "sister", "brother", "grandpa"],
    "health": ["doctor", "medical", "health", "therapy", "leg", "injury", "exercise", "weight"],
    "legal": ["lawyer", "legal", "court", "contract", "lawsuit", "attorney", "will", "estate"],
    "home": ["house", "home", "repair", "plumbing", "electric", "hvac", "mortgage", "renovation"],
    "auto": ["car", "toyota", "highlander", "vehicle", "mechanic", "tire", "oil"],
    "travel": ["travel", "flight", "hotel", "trip", "vacation", "passport", "airport"],
    "music": ["music", "song", "guitar", "band", "album", "spotify"],
    "business": ["business", "company", "startup", "entrepreneur", "marketing", "client", "consultant"],
    "writing": ["writing", "write", "story", "narrative", "voice", "author", "publisher"],
    "research": ["research", "compare", "analysis", "review", "best", "recommend", "options"],
}

# ChatGPT roles we import (skip system and tool)
IMPORTABLE_ROLES = {"user", "assistant"}

# Content types we extract text from
TEXT_CONTENT_TYPES = {"text", "multimodal_text", "code", "execution_output",
                      "thoughts", "reasoning_recap"}


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def green(text):
    return f"\033[92m{text}\033[0m"

def yellow(text):
    return f"\033[93m{text}\033[0m"

def red(text):
    return f"\033[91m{text}\033[0m"


def find_conversation_files(directory):
    """Find all conversations-NNN.json files in a directory."""
    d = Path(directory)
    files = sorted(d.glob("conversations-*.json"))
    if not files:
        # Maybe the zip was extracted flat -- check for conversations.json
        single = d / "conversations.json"
        if single.exists():
            files = [single]
    return files


def load_conversations(directory):
    """Load all conversations from JSON files in directory."""
    files = find_conversation_files(directory)
    if not files:
        print(red(f"No conversations-*.json files found in {directory}"))
        sys.exit(1)

    all_convos = []
    for f in files:
        with open(f, encoding="utf-8") as fh:
            convos = json.load(fh)
            all_convos.extend(convos)
        print(f"  Loaded {len(convos)} conversations from {f.name}")

    return all_convos


def extract_messages(conversation):
    """Extract messages from ChatGPT's tree-based mapping structure.

    Returns list of (role, content, timestamp) tuples in chronological order.
    Skips system messages, tool messages, and empty content.
    """
    mapping = conversation.get("mapping", {})
    if not mapping:
        return []

    # Build ordered message list by following parent→child chain
    messages = []
    for node_id, node in mapping.items():
        msg = node.get("message")
        if not msg:
            continue

        role = msg.get("author", {}).get("role", "")
        if role not in IMPORTABLE_ROLES:
            continue

        content_type = msg.get("content", {}).get("content_type", "")
        if content_type not in TEXT_CONTENT_TYPES:
            continue

        parts = msg.get("content", {}).get("parts", [])
        # Extract text from parts (skip dict parts which are images/files)
        text_parts = []
        for part in parts:
            if isinstance(part, str) and part.strip():
                text_parts.append(part.strip())
            elif isinstance(part, dict):
                # Some dict parts have text content
                if "text" in part:
                    text_parts.append(part["text"].strip())

        content = "\n".join(text_parts)
        if not content:
            continue

        timestamp = msg.get("create_time", 0)
        messages.append((role, content, timestamp, node_id))

    # Sort by timestamp
    messages.sort(key=lambda x: x[2])
    return messages


def suggest_project(title):
    """Auto-suggest a project based on conversation title keywords."""
    if not title:
        return DEFAULT_PROJECT

    title_lower = title.lower()
    for project, keywords in PROJECT_KEYWORDS.items():
        if any(kw in title_lower for kw in keywords):
            return project
    return DEFAULT_PROJECT


def suggest_tags(title, messages=None):
    """Auto-suggest tags based on title and optionally first few messages."""
    if not title:
        return ""

    search_text = title.lower()
    # Also check first 2 user messages for better tag coverage
    if messages:
        for role, content, _, _ in messages[:4]:
            if role == "user" and content:
                search_text += " " + content[:500].lower()

    matched_tags = []
    for tag, keywords in TAG_KEYWORDS.items():
        if any(kw in search_text for kw in keywords):
            matched_tags.append(tag)

    # Limit to 3 most relevant tags
    return ", ".join(matched_tags[:3]) if matched_tags else ""


def load_config():
    """Load config.yaml to get database path."""
    config_path = ROOT_DIR / "config.yaml"
    if not config_path.exists():
        print(red("config.yaml not found. Run brain-setup.py first."))
        sys.exit(1)

    import yaml
    with open(config_path) as f:
        config = yaml.safe_load(f)

    storage = config.get("storage", {})
    mode = storage.get("mode", "local")
    if mode == "synced":
        db_path = os.path.expanduser(storage.get("local_db_path", ""))
    else:
        db_path = os.path.join(os.path.expanduser(storage.get("root_path", "")), "claude-brain.db")

    return db_path


# ---------------------------------------------------------------------------
# SCAN mode: generate mapping CSV
# ---------------------------------------------------------------------------

def cmd_scan(args):
    """Scan ChatGPT export and generate mapping xlsx for review."""
    print(f"\nScanning ChatGPT export in: {args.directory}")
    print()

    convos = load_conversations(args.directory)

    # Check DB for already-imported conversations
    existing_ids = set()
    try:
        db_path = load_config()
        conn = sqlite3.connect(db_path)
        for row in conn.execute("SELECT session_id FROM sys_sessions WHERE source = 'chatgpt_import'"):
            existing_ids.add(row[0].replace("chatgpt_", ""))
        conn.close()
    except Exception:
        pass  # DB not available, all will show as NEW

    # Generate mapping
    rows = []
    for c in convos:
        title = c.get("title") or "(untitled)"
        conv_id = c.get("conversation_id", c.get("id", ""))
        create_time = c.get("create_time", 0)
        messages = extract_messages(c)
        msg_count = len(messages)
        date_str = datetime.fromtimestamp(create_time).strftime("%Y-%m-%d") if create_time else "unknown"
        suggested_project = suggest_project(title)
        suggested_tags = suggest_tags(title, messages)
        status = "IMPORTED" if conv_id in existing_ids else "NEW"

        rows.append({
            "conversation_id": conv_id,
            "date": date_str,
            "messages": msg_count,
            "project": suggested_project,
            "tags": suggested_tags,
            "status": status,
            "title": title,
        })

    # Sort by date
    rows.sort(key=lambda x: x["date"])

    # Write xlsx with two tabs
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    output_path = Path(args.output) if args.output else ROOT_DIR / "imports" / "chatgpt_import_map.xlsx"
    wb = Workbook()

    # --- Tab 1: Conversations ---
    ws1 = wb.active
    ws1.title = "Conversations"
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    edit_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    new_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    headers = ["conversation_id", "date", "messages", "project", "tags", "status", "title"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row_num, r in enumerate(rows, 2):
        ws1.cell(row=row_num, column=1, value=r["conversation_id"])
        ws1.cell(row=row_num, column=2, value=r["date"])
        ws1.cell(row=row_num, column=3, value=r["messages"])
        cell_proj = ws1.cell(row=row_num, column=4, value=r["project"])
        cell_proj.fill = edit_fill
        cell_tags = ws1.cell(row=row_num, column=5, value=r["tags"])
        cell_tags.fill = edit_fill
        cell_status = ws1.cell(row=row_num, column=6, value=r["status"])
        if r["status"] == "NEW":
            cell_status.fill = new_fill
        ws1.cell(row=row_num, column=7, value=r["title"])

    ws1.column_dimensions["A"].width = 40
    ws1.column_dimensions["B"].width = 12
    ws1.column_dimensions["C"].width = 10
    ws1.column_dimensions["D"].width = 10
    ws1.column_dimensions["E"].width = 30
    ws1.column_dimensions["F"].width = 12
    ws1.column_dimensions["G"].width = 60

    # --- Tab 2: Project Reference ---
    ws2 = wb.create_sheet("Project Reference")
    ws2.cell(row=1, column=1, value="USE THESE VALUES IN THE PROJECT COLUMN").font = Font(bold=True, size=14)
    ws2.merge_cells("A1:D1")

    ref_headers = ["prefix", "folder_name", "label", "status"]
    for col, h in enumerate(ref_headers, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    try:
        db_path = load_config()
        conn = sqlite3.connect(db_path)
        proj_rows = conn.execute("SELECT prefix, folder_name, label, status FROM project_registry ORDER BY folder_name").fetchall()
        for row_num, r in enumerate(proj_rows, 4):
            for col, val in enumerate(r, 1):
                ws2.cell(row=row_num, column=col, value=val)
        last_row = 4 + len(proj_rows)
        conn.close()
    except Exception:
        last_row = 4

    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 25
    ws2.column_dimensions["C"].width = 50
    ws2.column_dimensions["D"].width = 10

    ws2.cell(row=last_row + 1, column=1, value="").font = Font(bold=True)
    ws2.cell(row=last_row + 2, column=1, value="INSTRUCTIONS:").font = Font(bold=True)
    ws2.cell(row=last_row + 3, column=1, value="1. Go to the Conversations tab")
    ws2.cell(row=last_row + 4, column=1, value="2. Review the yellow PROJECT column -- change any wrong assignments")
    ws2.cell(row=last_row + 5, column=1, value="3. Review the yellow TAGS column -- edit, add, or remove tags")
    ws2.cell(row=last_row + 6, column=1, value="4. Delete entire rows for conversations you do NOT want imported")
    ws2.cell(row=last_row + 7, column=1, value="5. Rows marked IMPORTED will be skipped automatically")
    ws2.cell(row=last_row + 8, column=1, value="6. Save the file")
    ws2.cell(row=last_row + 9, column=1, value="7. Tell Claude you are ready for dry-run")

    wb.save(output_path)

    # Summary
    project_counts = {}
    tag_counts = {}
    total_msgs = 0
    new_count = sum(1 for r in rows if r["status"] == "NEW")
    for r in rows:
        p = r["project"]
        project_counts[p] = project_counts.get(p, 0) + 1
        total_msgs += r["messages"]
        for tag in [t.strip() for t in r["tags"].split(",") if t.strip()]:
            tag_counts[tag] = tag_counts.get(tag, 0) + 1

    print(f"\n{green('Mapping xlsx generated:')} {output_path}")
    print(f"\n  Conversations: {len(rows)} ({new_count} NEW, {len(rows) - new_count} IMPORTED)")
    print(f"  Total messages: {total_msgs}")
    print(f"\n  Project suggestions:")
    for p, count in sorted(project_counts.items()):
        print(f"    {p}: {count} conversations")
    print(f"\n  Tag suggestions ({len(tag_counts)} unique tags):")
    for tag, count in sorted(tag_counts.items(), key=lambda x: -x[1]):
        print(f"    {tag}: {count} conversations")

    print(f"\n{yellow('NEXT STEP:')} Review and edit {output_path.name}")
    print(f"  - Edit the yellow PROJECT and TAGS columns")
    print(f"  - Delete rows you don't want imported")
    print(f"  - Then run: python3 scripts/import_chatgpt.py --import {args.directory} --map {output_path}")


# ---------------------------------------------------------------------------
# IMPORT mode: read mapping CSV and ingest
# ---------------------------------------------------------------------------

def cmd_import(args):
    """Import conversations using reviewed mapping CSV."""
    if not args.map:
        print(red("--map <csv_file> is required for import mode"))
        sys.exit(1)

    map_path = Path(args.map)
    if not map_path.exists():
        print(red(f"Mapping file not found: {map_path}"))
        print(f"Run --scan first to generate it.")
        sys.exit(1)

    # Read mapping from xlsx or csv
    mapping = {}  # conv_id -> {"project": ..., "tags": ...}
    if map_path.suffix == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(map_path)
        ws = wb["Conversations"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        for row in range(2, ws.max_row + 1):
            conv_id = ws.cell(row=row, column=1).value
            if not conv_id:
                continue
            project = ws.cell(row=row, column=headers.index("project") + 1).value or DEFAULT_PROJECT
            tags = ws.cell(row=row, column=headers.index("tags") + 1).value or ""
            mapping[conv_id] = {"project": project, "tags": tags}
    else:
        with open(map_path, encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                mapping[row["conversation_id"]] = {
                    "project": row.get("project", DEFAULT_PROJECT),
                    "tags": row.get("tags", ""),
                }

    print(f"\nMapping file: {map_path.name} ({len(mapping)} conversations)")

    # Load conversations
    print(f"Loading conversations from: {args.directory}")
    convos = load_conversations(args.directory)

    # Match conversations to mapping
    to_import = []
    skipped_not_in_map = 0
    skipped_zero_msgs = 0

    for c in convos:
        conv_id = c.get("conversation_id", c.get("id", ""))
        if conv_id not in mapping:
            skipped_not_in_map += 1
            continue

        messages = extract_messages(c)
        if not messages:
            skipped_zero_msgs += 1
            continue

        to_import.append({
            "conversation": c,
            "messages": messages,
            "project": mapping[conv_id]["project"],
            "tags": mapping[conv_id].get("tags", ""),
            "conv_id": conv_id,
        })

    print(f"\n  To import: {len(to_import)} conversations")
    print(f"  Skipped (not in mapping): {skipped_not_in_map}")
    print(f"  Skipped (0 messages): {skipped_zero_msgs}")

    if args.dry_run:
        print(f"\n{yellow('DRY RUN')} -- showing what would be imported:\n")
        project_summary = {}
        for item in to_import:
            p = item["project"]
            if p not in project_summary:
                project_summary[p] = {"count": 0, "messages": 0}
            project_summary[p]["count"] += 1
            project_summary[p]["messages"] += len(item["messages"])

        print(f"  {'Project':<12} {'Conversations':>14} {'Messages':>10}")
        print(f"  {'-'*12} {'-'*14} {'-'*10}")
        for p in sorted(project_summary.keys()):
            s = project_summary[p]
            print(f"  {p:<12} {s['count']:>14} {s['messages']:>10}")
        total_msgs = sum(s["messages"] for s in project_summary.values())
        print(f"  {'TOTAL':<12} {len(to_import):>14} {total_msgs:>10}")

        print(f"\n  Sample conversations:")
        for item in to_import[:10]:
            c = item["conversation"]
            title = (c.get("title") or "(untitled)")[:60]
            print(f"    [{item['project']}] {len(item['messages']):>3} msgs | {title}")
        if len(to_import) > 10:
            print(f"    ... and {len(to_import) - 10} more")

        print(f"\n{yellow('To execute:')} remove --dry-run flag")
        return

    # Actually import
    db_path = load_config()
    conn = sqlite3.connect(db_path)
    now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    imported = 0
    skipped_dup = 0
    total_transcripts = 0

    for item in to_import:
        c = item["conversation"]
        conv_id = item["conv_id"]
        project = item["project"]
        messages = item["messages"]
        title = c.get("title", "") or "(untitled)"
        create_time = c.get("create_time", 0)

        # Check for duplicate (by conversation_id in session_id)
        existing = conn.execute(
            "SELECT 1 FROM sys_sessions WHERE session_id = ?",
            (f"chatgpt_{conv_id}",)
        ).fetchone()

        if existing:
            skipped_dup += 1
            continue

        # Create session
        tags = item.get("tags", "")
        started = datetime.fromtimestamp(create_time).strftime("%Y-%m-%dT%H:%M:%SZ") if create_time else now
        ended = datetime.fromtimestamp(messages[-1][2]).strftime("%Y-%m-%dT%H:%M:%SZ") if messages[-1][2] else now

        conn.execute(
            """INSERT INTO sys_sessions
               (session_id, project, started_at, ended_at, cwd, claude_version, model, source, message_count, created_at, notes, tags)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (f"chatgpt_{conv_id}", project, started, ended,
             "chatgpt", "chatgpt", "gpt", "chatgpt_import",
             len(messages), now, f"ChatGPT import: {title}", tags)
        )

        # Insert messages
        for i, (role, content, timestamp, node_id) in enumerate(messages):
            ts = datetime.fromtimestamp(timestamp).strftime("%Y-%m-%dT%H:%M:%SZ") if timestamp else now
            sender = "human" if role == "user" else "assistant"

            conn.execute(
                """INSERT INTO transcripts
                   (session_id, project, uuid, type, role, content, timestamp, source, created_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (f"chatgpt_{conv_id}", project, node_id,
                 "text", sender, content, ts, "chatgpt", now)
            )
            total_transcripts += 1

        imported += 1

    conn.commit()
    conn.close()

    print(f"\n{green('Import complete:')}")
    print(f"  Conversations imported: {imported}")
    print(f"  Transcripts created: {total_transcripts}")
    print(f"  Skipped (already imported): {skipped_dup}")
    print(f"  Source tag: 'chatgpt' (transcripts), 'chatgpt_import' (sessions)")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Import ChatGPT conversations into the brain database"
    )
    parser.add_argument("--scan", dest="scan_dir", metavar="DIR",
                        help="Scan ChatGPT export directory and generate mapping CSV")
    parser.add_argument("--import", dest="import_dir", metavar="DIR",
                        help="Import conversations using mapping CSV")
    parser.add_argument("--map", metavar="CSV",
                        help="Mapping CSV file (required for --import)")
    parser.add_argument("--output", metavar="FILE",
                        help="Output path for mapping CSV (default: chatgpt_import_map.csv)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Show what would be imported without touching the database")

    args = parser.parse_args()

    if args.scan_dir:
        args.directory = args.scan_dir
        cmd_scan(args)
    elif args.import_dir:
        args.directory = args.import_dir
        cmd_import(args)
    else:
        parser.print_help()
        print(f"\n{yellow('Workflow:')}")
        print(f"  1. python3 scripts/import_chatgpt.py --scan <chatgpt-export-dir>")
        print(f"  2. Review chatgpt_import_map.csv (edit project assignments)")
        print(f"  3. python3 scripts/import_chatgpt.py --import <dir> --map chatgpt_import_map.csv --dry-run")
        print(f"  4. python3 scripts/import_chatgpt.py --import <dir> --map chatgpt_import_map.csv")


if __name__ == "__main__":
    main()
