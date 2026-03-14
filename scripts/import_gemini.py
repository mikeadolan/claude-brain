#!/usr/bin/env python3
"""
import_gemini.py - Import Google Gemini conversation exports into the brain database.

Supports Google Takeout exports (My Activity > Gemini Apps > MyActivity.html).
Text conversations only -- images, audio, and attachments are skipped.

Three modes:
  --scan <dir>     Generate mapping xlsx with suggested project/tag assignments
  --import <dir>   Import conversations using reviewed mapping xlsx
  --dry-run        Preview import without touching the database

Session grouping: Gemini exports individual exchanges, not sessions.
Exchanges within 30 minutes of each other are grouped into one session.

Workflow:
  1. Export from Google Takeout (My Activity > Gemini Apps)
  2. Extract the zip to a folder
  3. Run: python3 scripts/import_gemini.py --scan imports/gemini-export/
  4. Review gemini_import_map.xlsx -- edit project/tag assignments, delete unwanted rows
  5. Run: python3 scripts/import_gemini.py --import imports/gemini-export/ --map gemini_import_map.xlsx --dry-run
  6. When satisfied: python3 scripts/import_gemini.py --import imports/gemini-export/ --map gemini_import_map.xlsx
"""

import argparse
import csv
import json
import os
import re
import sqlite3
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
ROOT_DIR = SCRIPT_DIR.parent

# Session grouping: exchanges within this many minutes = same session
SESSION_GAP_MINUTES = 30

# Project keyword suggestions (same as ChatGPT import)
PROJECT_KEYWORDS = {
    "jg": ["johnny", "goods", "memoir", "chapter", "ch1", "ch2", "ch3", "ch4",
            "ch5", "ch6", "ch7", "ch8", "ch9", "ch10", "ch11", "polish",
            "manuscript", "mob", "harlem", "mafia", "maffia", "three fingers",
            "gangster", "dialogue", "copyedit", "proofread", "book edit",
            "edit instructions", "summary deliverable", "pleasant avenue",
            "east harlem", "john maffia"],
    "js": ["resume", "job", "interview", "career", "hiring", "salary",
           "linkedin", "cover letter", "recruiter"],
}
DEFAULT_PROJECT = "gen"

# Tag keyword suggestions (same as ChatGPT import)
TAG_KEYWORDS = {
    "book-editing": ["edit", "polish", "chapter", "manuscript", "copyedit", "proofread", "rewrite", "revision"],
    "memoir": ["memoir", "johnny", "goods", "maffia", "mob", "harlem", "gangster", "three fingers", "pleasant avenue"],
    "dialogue": ["dialogue", "dialog", "conversation", "speech", "quotes"],
    "job-search": ["resume", "job", "interview", "career", "hiring", "salary", "recruiter", "linkedin"],
    "finance": ["money", "bank", "invest", "stock", "crypto", "budget", "tax", "financial"],
    "coding": ["python", "script", "code", "programming", "debug", "function", "api", "github"],
    "ai-tools": ["chatgpt", "claude", "gpt", "openai", "anthropic", "gemini", "ai model", "llm", "prompt"],
    "tech-setup": ["linux", "fedora", "install", "setup", "config", "terminal", "laptop", "asus", "ubuntu", "samba"],
    "family": ["mom", "mother", "father", "wife", "daughter", "son", "family", "sister", "brother", "grandpa"],
    "health": ["doctor", "medical", "health", "therapy", "leg", "injury", "exercise", "weight"],
    "legal": ["lawyer", "legal", "court", "contract", "lawsuit", "attorney", "will", "estate"],
    "home": ["house", "home", "repair", "plumbing", "electric", "hvac", "mortgage", "renovation"],
    "auto": ["car", "toyota", "highlander", "vehicle", "mechanic", "tire", "oil"],
    "travel": ["travel", "flight", "hotel", "trip", "vacation", "passport", "airport"],
    "music": ["music", "song", "guitar", "band", "album", "spotify"],
    "business": ["business", "company", "startup", "entrepreneur", "marketing", "client", "consultant"],
    "writing": ["writing", "write", "story", "narrative", "voice", "author", "publisher"],
    "research": ["research", "compare", "analysis", "review", "best", "recommend", "options"],
}


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def green(text):
    return f"\033[92m{text}\033[0m"

def yellow(text):
    return f"\033[93m{text}\033[0m"

def red(text):
    return f"\033[91m{text}\033[0m"


def load_config():
    """Load config.yaml to get database path."""
    config_path = ROOT_DIR / "config.yaml"
    if not config_path.exists():
        print(red("config.yaml not found. Run brain-setup.py first."))
        sys.exit(1)

    import yaml
    with open(config_path) as f:
        config = yaml.safe_load(f)

    storage = config.get("storage", {})
    mode = storage.get("mode", "local")
    if mode == "synced":
        db_path = os.path.expanduser(storage.get("local_db_path", ""))
    else:
        db_path = os.path.join(os.path.expanduser(storage.get("root_path", "")), "claude-brain.db")

    return db_path


def suggest_project(text):
    """Auto-suggest a project based on text keywords."""
    if not text:
        return DEFAULT_PROJECT
    text_lower = text.lower()
    for project, keywords in PROJECT_KEYWORDS.items():
        if any(kw in text_lower for kw in keywords):
            return project
    return DEFAULT_PROJECT


def suggest_tags(text):
    """Auto-suggest tags based on text content."""
    if not text:
        return ""
    text_lower = text.lower()
    matched = []
    for tag, keywords in TAG_KEYWORDS.items():
        if any(kw in text_lower for kw in keywords):
            matched.append(tag)
    return ", ".join(matched[:3]) if matched else ""


def generate_title(session_exchanges):
    """Generate a meaningful title from the session's prompt content.

    Takes the first user prompt and extracts a clean, short title.
    Handles: 'Prompted ...', 'Created Gemini Canvas titled ...',
    'Used an Assistant feature', and raw prompts.
    """
    # Combine first 2-3 prompts for context
    prompts = []
    for prompt, response, ts in session_exchanges[:3]:
        if prompt:
            prompts.append(prompt)

    if not prompts:
        return "(no content)"

    text = prompts[0]

    # Clean prefixes
    if text.startswith("Prompted "):
        text = text[9:]
    if text.startswith("Created Gemini Canvas titled "):
        text = text[28:]
    if text.startswith("Used an Assistant feature") or not text.strip():
        # No text content -- try second or third prompt in session
        found = False
        for p in prompts[1:]:
            clean = p
            if clean.startswith("Prompted "):
                clean = clean[9:]
            if clean.startswith("Used an Assistant feature"):
                continue
            if clean.strip():
                text = clean
                found = True
                break
        if not found:
            # Try response content as last resort
            for _, response, _ in session_exchanges[:3]:
                if response and len(response) > 20:
                    text = response
                    found = True
                    break
        if not found:
            text = "(no text content - voice/image only)"

    # Remove file attachment references
    text = re.sub(r'Attached \d+ file\(s?\)\..*$', '', text, flags=re.DOTALL)
    text = re.sub(r'-\s*image_\w+\.\w+', '', text)

    # Remove URLs
    text = re.sub(r'http\S+', '', text)

    # Take first sentence or first 80 chars
    text = text.strip()
    # Split on sentence boundaries
    sentence_end = re.search(r'[.!?\n]', text)
    if sentence_end and sentence_end.start() > 10:
        text = text[:sentence_end.start()]

    # Truncate
    if len(text) > 80:
        # Try to break at a word boundary
        text = text[:80]
        last_space = text.rfind(' ')
        if last_space > 40:
            text = text[:last_space]

    return text.strip() or "(no content)"


# ---------------------------------------------------------------------------
# HTML Parsing
# ---------------------------------------------------------------------------

def find_html_file(directory):
    """Find MyActivity.html in the directory (may be nested in Takeout/)."""
    d = Path(directory)
    # Direct
    if (d / "MyActivity.html").exists():
        return d / "MyActivity.html"
    # Nested in Takeout structure
    nested = d / "Takeout" / "My Activity" / "Gemini Apps" / "MyActivity.html"
    if nested.exists():
        return nested
    # Search recursively
    for f in d.rglob("MyActivity.html"):
        return f
    return None


def parse_exchanges(html_path):
    """Parse Gemini MyActivity.html into individual exchanges.

    Returns list of (prompt, response, timestamp_dt) tuples.
    """
    with open(html_path, encoding="utf-8") as f:
        html = f.read()

    # Split by outer-cell blocks
    blocks = re.split(r'<div class="outer-cell[^"]*">', html)
    blocks = [b for b in blocks[1:] if b.strip()]

    exchanges = []
    for block in blocks:
        # Extract the main content cell
        match = re.search(
            r'mdl-typography--body-1">(.*?)</div>\s*<div class="content-cell[^"]*'
            r'mdl-typography--body-1 mdl-typography--text-right',
            block, re.DOTALL
        )
        if not match:
            # Try simpler pattern
            match = re.search(r'mdl-typography--body-1">(.*?)</div>', block, re.DOTALL)
        if not match:
            continue

        full_content = match.group(1)

        # Find timestamp (handles unicode narrow no-break space \u202f before AM/PM)
        ts_match = re.search(
            r'(\w{3} \d{1,2}, \d{4}, \d{1,2}:\d{2}:\d{2})\s*\u202f?\s*([AP]M)\s*(\w+)',
            full_content
        )
        if not ts_match:
            continue

        # Parse timestamp
        try:
            dt = datetime.strptime(
                f'{ts_match.group(1)} {ts_match.group(2)}',
                '%b %d, %Y, %I:%M:%S %p'
            )
        except ValueError:
            continue

        # Split content on timestamp
        before_ts = full_content[:ts_match.start()]
        after_ts = full_content[ts_match.end():]

        # Extract prompt (before timestamp)
        prompt = re.sub(r'<[^>]+>', '', before_ts).strip()
        # Remove "Prompted " prefix
        if prompt.startswith("Prompted "):
            prompt = prompt[9:]
        elif prompt.startswith("Created Gemini Canvas titled "):
            prompt = prompt  # keep the full text

        # Extract response (after timestamp)
        response = re.sub(r'<[^>]+>', ' ', after_ts).strip()
        response = re.sub(r'\s+', ' ', response)
        # Clean HTML entities
        response = response.replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>')
        response = response.replace('&quot;', '"').replace('&#39;', "'")

        if prompt or response:
            exchanges.append((prompt.strip(), response.strip(), dt))

    # Reverse to chronological order (HTML is newest-first)
    exchanges.reverse()
    return exchanges


def group_into_sessions(exchanges, gap_minutes=SESSION_GAP_MINUTES):
    """Group exchanges into sessions based on time proximity.

    Exchanges within gap_minutes of each other belong to the same session.
    Returns list of sessions, each a list of (prompt, response, timestamp) tuples.
    """
    if not exchanges:
        return []

    sessions = []
    current_session = [exchanges[0]]

    for i in range(1, len(exchanges)):
        gap = (exchanges[i][2] - exchanges[i - 1][2]).total_seconds() / 60
        if gap > gap_minutes:
            sessions.append(current_session)
            current_session = [exchanges[i]]
        else:
            current_session.append(exchanges[i])

    if current_session:
        sessions.append(current_session)

    return sessions


# ---------------------------------------------------------------------------
# SCAN mode
# ---------------------------------------------------------------------------

def cmd_scan(args):
    """Scan Gemini export and generate mapping xlsx for review."""
    print(f"\nScanning Gemini export in: {args.directory}")

    html_file = find_html_file(args.directory)
    if not html_file:
        print(red("MyActivity.html not found. Make sure the Takeout zip is extracted."))
        sys.exit(1)

    print(f"  Found: {html_file}")
    exchanges = parse_exchanges(html_file)
    print(f"  Parsed: {len(exchanges)} exchanges")

    sessions = group_into_sessions(exchanges)
    print(f"  Grouped into: {len(sessions)} sessions ({SESSION_GAP_MINUTES}-min gap)")

    # Check DB for already-imported sessions
    existing_ids = set()
    try:
        db_path = load_config()
        conn = sqlite3.connect(db_path)
        for row in conn.execute("SELECT session_id FROM sys_sessions WHERE source = 'gemini_import'"):
            existing_ids.add(row[0])
        conn.close()
    except Exception:
        pass

    # Build session metadata
    rows = []
    for i, session in enumerate(sessions):
        first_ts = session[0][2]
        last_ts = session[-1][2]
        msg_count = len(session) * 2  # each exchange = 1 user + 1 assistant

        # Generate session ID based on first timestamp
        session_id = f"gemini_{first_ts.strftime('%Y%m%d_%H%M%S')}"

        # Build combined text from ALL prompts + responses for better suggestions
        combined_text = " ".join(
            (ex[0][:300] + " " + ex[1][:300]) for ex in session
        )

        # Generate meaningful title from content
        title = generate_title(session)

        suggested_project = suggest_project(combined_text)
        suggested_tags = suggest_tags(combined_text)
        status = "IMPORTED" if session_id in existing_ids else "NEW"

        rows.append({
            "session_id": session_id,
            "date": first_ts.strftime("%Y-%m-%d"),
            "time": first_ts.strftime("%H:%M"),
            "exchanges": len(session),
            "messages": msg_count,
            "project": suggested_project,
            "tags": suggested_tags,
            "status": status,
            "title": title,
        })

    # Write xlsx
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    output_path = Path(args.output) if args.output else ROOT_DIR / "imports" / "gemini_import_map.xlsx"
    wb = Workbook()

    # Tab 1: Sessions
    ws1 = wb.active
    ws1.title = "Sessions"
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    edit_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    new_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    headers = ["session_id", "date", "time", "exchanges", "messages", "project", "tags", "status", "title"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row_num, r in enumerate(rows, 2):
        ws1.cell(row=row_num, column=1, value=r["session_id"])
        ws1.cell(row=row_num, column=2, value=r["date"])
        ws1.cell(row=row_num, column=3, value=r["time"])
        ws1.cell(row=row_num, column=4, value=r["exchanges"])
        ws1.cell(row=row_num, column=5, value=r["messages"])
        cell_proj = ws1.cell(row=row_num, column=6, value=r["project"])
        cell_proj.fill = edit_fill
        cell_tags = ws1.cell(row=row_num, column=7, value=r["tags"])
        cell_tags.fill = edit_fill
        cell_status = ws1.cell(row=row_num, column=8, value=r["status"])
        if r["status"] == "NEW":
            cell_status.fill = new_fill
        ws1.cell(row=row_num, column=9, value=r["title"])

    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 12
    ws1.column_dimensions["C"].width = 8
    ws1.column_dimensions["D"].width = 12
    ws1.column_dimensions["E"].width = 10
    ws1.column_dimensions["F"].width = 10
    ws1.column_dimensions["G"].width = 30
    ws1.column_dimensions["H"].width = 12
    ws1.column_dimensions["I"].width = 60

    # Tab 2: Project Reference
    ws2 = wb.create_sheet("Project Reference")
    ws2.cell(row=1, column=1, value="USE THESE VALUES IN THE PROJECT COLUMN").font = Font(bold=True, size=14)
    ws2.merge_cells("A1:D1")
    ref_headers = ["prefix", "folder_name", "label", "status"]
    for col, h in enumerate(ref_headers, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
    try:
        db_path = load_config()
        conn = sqlite3.connect(db_path)
        proj_rows = conn.execute("SELECT prefix, folder_name, label, status FROM project_registry ORDER BY folder_name").fetchall()
        for row_num, r in enumerate(proj_rows, 4):
            for col, val in enumerate(r, 1):
                ws2.cell(row=row_num, column=col, value=val)
        last_row = 4 + len(proj_rows)
        conn.close()
    except Exception:
        last_row = 4
    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 25
    ws2.column_dimensions["C"].width = 50
    ws2.column_dimensions["D"].width = 10

    # Tab 3: Tag Reference
    ws3 = wb.create_sheet("Tag Reference")
    ws3.cell(row=1, column=1, value="AVAILABLE TAGS").font = Font(bold=True, size=14)
    tag_headers = ["tag", "sessions", "description"]
    for col, h in enumerate(tag_headers, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
    tag_counts = {}
    for r in rows:
        for tag in [t.strip() for t in r["tags"].split(",") if t.strip()]:
            tag_counts[tag] = tag_counts.get(tag, 0) + 1
    tag_descriptions = {
        "auto": "Cars, vehicles, maintenance", "ai-tools": "ChatGPT, Claude, Gemini, LLMs",
        "book-editing": "Chapter edits, polish, revisions", "business": "Companies, startups, marketing",
        "coding": "Python, scripts, programming", "dialogue": "Dialogue writing in the book",
        "family": "Mom, father, wife, kids, grandpa", "finance": "Money, banking, investing, taxes",
        "health": "Medical, therapy, injuries", "home": "House repairs, plumbing, electrical",
        "job-search": "Resumes, interviews, career", "legal": "Lawyers, courts, contracts, estates",
        "memoir": "Johnny Goods, mob stories, Harlem", "music": "Songs, guitar, music project",
        "research": "Comparisons, analysis, reviews", "tech-setup": "Linux, Fedora, installs, configs",
        "travel": "Flights, hotels, trips", "writing": "Narrative craft, voice, storytelling",
    }
    for row_num, (tag, count) in enumerate(sorted(tag_counts.items(), key=lambda x: -x[1]), 4):
        ws3.cell(row=row_num, column=1, value=tag)
        ws3.cell(row=row_num, column=2, value=count)
        ws3.cell(row=row_num, column=3, value=tag_descriptions.get(tag, ""))
    ws3.column_dimensions["A"].width = 18
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 50

    wb.save(output_path)

    # Summary
    project_counts = {}
    total_msgs = 0
    new_count = sum(1 for r in rows if r["status"] == "NEW")
    for r in rows:
        project_counts[r["project"]] = project_counts.get(r["project"], 0) + 1
        total_msgs += r["messages"]

    print(f"\n{green('Mapping xlsx generated:')} {output_path}")
    print(f"\n  Sessions: {len(rows)} ({new_count} NEW, {len(rows) - new_count} IMPORTED)")
    print(f"  Total messages: {total_msgs}")
    print(f"  Exchanges: {len(exchanges)}")
    print(f"\n  Project suggestions:")
    for p, count in sorted(project_counts.items()):
        print(f"    {p}: {count} sessions")
    print(f"\n  Tags: {len(tag_counts)} unique")
    for tag, count in sorted(tag_counts.items(), key=lambda x: -x[1])[:10]:
        print(f"    {tag}: {count}")

    print(f"\n{yellow('NEXT STEP:')} Review and edit {output_path.name}")


# ---------------------------------------------------------------------------
# IMPORT mode
# ---------------------------------------------------------------------------

def cmd_import(args):
    """Import sessions using reviewed mapping xlsx."""
    if not args.map:
        print(red("--map <xlsx_file> is required for import mode"))
        sys.exit(1)

    map_path = Path(args.map)
    if not map_path.exists():
        print(red(f"Mapping file not found: {map_path}"))
        sys.exit(1)

    # Read mapping
    from openpyxl import load_workbook
    wb = load_workbook(map_path)
    ws = wb["Sessions"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    mapping = {}
    for row in range(2, ws.max_row + 1):
        sid = ws.cell(row=row, column=1).value
        if not sid:
            continue
        mapping[sid] = {
            "project": ws.cell(row=row, column=headers.index("project") + 1).value or DEFAULT_PROJECT,
            "tags": ws.cell(row=row, column=headers.index("tags") + 1).value or "",
        }

    print(f"\nMapping file: {map_path.name} ({len(mapping)} sessions)")

    # Parse HTML
    html_file = find_html_file(args.directory)
    if not html_file:
        print(red("MyActivity.html not found."))
        sys.exit(1)

    print(f"Parsing: {html_file}")
    exchanges = parse_exchanges(html_file)
    sessions = group_into_sessions(exchanges)
    print(f"  {len(exchanges)} exchanges → {len(sessions)} sessions")

    # Match sessions to mapping
    to_import = []
    skipped_not_in_map = 0

    for session in sessions:
        first_ts = session[0][2]
        session_id = f"gemini_{first_ts.strftime('%Y%m%d_%H%M%S')}"

        if session_id not in mapping:
            skipped_not_in_map += 1
            continue

        to_import.append({
            "session_id": session_id,
            "exchanges": session,
            "project": mapping[session_id]["project"],
            "tags": mapping[session_id]["tags"],
        })

    print(f"\n  To import: {len(to_import)} sessions")
    print(f"  Skipped (not in mapping): {skipped_not_in_map}")

    if args.dry_run:
        print(f"\n{yellow('DRY RUN')} -- showing what would be imported:\n")
        project_summary = {}
        for item in to_import:
            p = item["project"]
            if p not in project_summary:
                project_summary[p] = {"count": 0, "messages": 0}
            project_summary[p]["count"] += 1
            project_summary[p]["messages"] += len(item["exchanges"]) * 2

        print(f"  {'Project':<12} {'Sessions':>10} {'Messages':>10}")
        print(f"  {'-'*12} {'-'*10} {'-'*10}")
        for p in sorted(project_summary.keys()):
            s = project_summary[p]
            print(f"  {p:<12} {s['count']:>10} {s['messages']:>10}")
        total_msgs = sum(s["messages"] for s in project_summary.values())
        print(f"  {'TOTAL':<12} {len(to_import):>10} {total_msgs:>10}")

        print(f"\n  Sample sessions:")
        for item in to_import[:10]:
            title = item["exchanges"][0][0][:55] if item["exchanges"][0][0] else "(no prompt)"
            print(f"    [{item['project']}] {len(item['exchanges']):>3} exch | {title}")

        print(f"\n{yellow('To execute:')} remove --dry-run flag")
        return

    # Actually import
    db_path = load_config()
    conn = sqlite3.connect(db_path)
    now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    imported = 0
    skipped_dup = 0
    total_transcripts = 0

    for item in to_import:
        session_id = item["session_id"]
        project = item["project"]
        tags = item["tags"]
        exchanges = item["exchanges"]
        title = (exchanges[0][0] or "(no prompt)")[:80]

        # Check for duplicate
        existing = conn.execute(
            "SELECT 1 FROM sys_sessions WHERE session_id = ?", (session_id,)
        ).fetchone()
        if existing:
            skipped_dup += 1
            continue

        # Create session
        started = exchanges[0][2].strftime("%Y-%m-%dT%H:%M:%SZ")
        ended = exchanges[-1][2].strftime("%Y-%m-%dT%H:%M:%SZ")

        conn.execute(
            """INSERT INTO sys_sessions
               (session_id, project, started_at, ended_at, cwd, claude_version, model, source,
                message_count, created_at, notes, tags)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (session_id, project, started, ended,
             "gemini", "gemini", "gemini", "gemini_import",
             len(exchanges) * 2, now,
             f"Gemini import: {title} ({len(exchanges)} exchanges)", tags)
        )

        # Insert messages (each exchange = 1 user + 1 assistant)
        for ex_idx, (prompt, response, ts) in enumerate(exchanges):
            ts_str = ts.strftime("%Y-%m-%dT%H:%M:%SZ")
            exchange_id = f"{session_id}_{ex_idx:04d}"

            if prompt:
                conn.execute(
                    """INSERT INTO transcripts
                       (session_id, project, uuid, type, role, content, timestamp, source, created_at)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (session_id, project, f"{exchange_id}_user",
                     "text", "human", prompt, ts_str, "gemini", now)
                )
                total_transcripts += 1

            if response:
                conn.execute(
                    """INSERT INTO transcripts
                       (session_id, project, uuid, type, role, content, timestamp, source, created_at)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (session_id, project, f"{exchange_id}_asst",
                     "text", "assistant", response, ts_str, "gemini", now)
                )
                total_transcripts += 1

        imported += 1

    conn.commit()
    conn.close()

    print(f"\n{green('Import complete:')}")
    print(f"  Sessions imported: {imported}")
    print(f"  Transcripts created: {total_transcripts}")
    print(f"  Skipped (already imported): {skipped_dup}")
    print(f"  Source tag: 'gemini' (transcripts), 'gemini_import' (sessions)")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Import Gemini conversations into the brain database"
    )
    parser.add_argument("--scan", dest="scan_dir", metavar="DIR",
                        help="Scan Gemini export directory and generate mapping xlsx")
    parser.add_argument("--import", dest="import_dir", metavar="DIR",
                        help="Import sessions using mapping xlsx")
    parser.add_argument("--map", metavar="XLSX",
                        help="Mapping xlsx file (required for --import)")
    parser.add_argument("--output", metavar="FILE",
                        help="Output path for mapping xlsx")
    parser.add_argument("--dry-run", action="store_true",
                        help="Show what would be imported without touching the database")

    args = parser.parse_args()

    if args.scan_dir:
        args.directory = args.scan_dir
        cmd_scan(args)
    elif args.import_dir:
        args.directory = args.import_dir
        cmd_import(args)
    else:
        parser.print_help()
        print(f"\n{yellow('Workflow:')}")
        print(f"  1. python3 scripts/import_gemini.py --scan <gemini-export-dir>")
        print(f"  2. Review gemini_import_map.xlsx (edit project/tag assignments)")
        print(f"  3. python3 scripts/import_gemini.py --import <dir> --map gemini_import_map.xlsx --dry-run")
        print(f"  4. python3 scripts/import_gemini.py --import <dir> --map gemini_import_map.xlsx")


if __name__ == "__main__":
    main()
